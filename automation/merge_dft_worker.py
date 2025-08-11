# automation/merge_dft_worker.py
from PySide6.QtCore import QThread, Signal
import os, time, shutil, csv
import pandas as pd
from openpyxl import load_workbook
import re

class MergeDftWorker(QThread):
    """
    Merge Excel + CSV pairs into {base}_merged.xlsx at the top folder,
    then move *everything else* (all originals including .iPrd/.csv/Excel and all subfolders)
    into backup/, preserving relative structure. Only merged files remain in the top folder.

    NOTE: This version appends CSV rows using openpyxl/csv.reader,
    so values in the 'DFT result' sheet are written as strings (pre-numeric-coercion behavior).
    """
    # Signals
    progress_updated = Signal(int, int, str)      # percent, current_index(1-based), base_name
    status_updated   = Signal(str)                # freeform text
    log_message      = Signal(str)                # freeform text (append to log window)
    finished         = Signal(list, list)         # success_list [(merged, xfile, cfile)], error_list [(base, err)]

    def __init__(self, top_folder: str):
        super().__init__()
        self.top_folder = os.path.abspath(top_folder)
        self._is_cancelled = False

    def cancel(self):
        self._is_cancelled = True

    # ---------------- helpers ----------------
    def _fmt_rel(self, path: str) -> str:
        try:
            return os.path.relpath(path, self.top_folder)
        except Exception:
            return path

    def _ensure_parent_dir(self, dst_path: str):
        os.makedirs(os.path.dirname(dst_path), exist_ok=True)

    def _safe_move(self, src: str, dst: str):
        self._ensure_parent_dir(dst)
        if os.path.abspath(src) == os.path.abspath(dst):
            return
        base, ext = os.path.splitext(dst)
        candidate = dst
        n = 1
        while os.path.exists(candidate):
            candidate = f"{base} ({n}){ext}"
            n += 1
        shutil.move(src, candidate)
        self.log_message.emit(f"↪️  moved: {self._fmt_rel(src)}  →  {self._fmt_rel(candidate)}")

    def _merge_pair(self, xfile: str, cfile: str, merged_fp: str):
        # 1) Copy all sheets from Excel into merged workbook
        all_sheets = pd.read_excel(xfile, sheet_name=None)
        with pd.ExcelWriter(merged_fp, engine='openpyxl') as writer:
            for sheet_name, df_sheet in all_sheets.items():
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

        # 2) Append raw CSV content as "DFT result" *as strings*
        raw = open(cfile, 'rb').read().decode('utf-8', errors='replace')
        wb = load_workbook(merged_fp)
        ws = wb.create_sheet("DFT result")
        for row in csv.reader(raw.splitlines()):
            ws.append(row)
        wb.save(merged_fp)

        self._fix_dft_sheet_numbers(merged_fp, sheet_name="DFT result")
    # ---------------- main work ----------------
    def run(self):
        t0 = time.time()
        try:
            self.status_updated.emit(f"Top: {self.top_folder}")
            backup_dir = os.path.join(self.top_folder, "backup")
            self.status_updated.emit(f"Backup: {backup_dir}")
            os.makedirs(backup_dir, exist_ok=True)
            self.log_message.emit(f"📁 ensure backup exists: {self._fmt_rel(backup_dir)}")
        except Exception as e:
            self.log_message.emit(f"❌ cannot create backup dir: {e}")
            self.finished.emit([], [("ALL", f"cannot create backup dir: {e}")])
            return

        excel_exts = ('.xls', '.xlsx', '.xlsm', '.xlsb')

        # --- gather candidates (include .iprd too) ---
        all_paths = []
        for dirpath, _, filenames in os.walk(self.top_folder):
            abs_dir = os.path.abspath(dirpath)
            # skip backup subtree entirely
            if abs_dir.startswith(os.path.abspath(backup_dir)):
                continue
            for fn in filenames:
                lower = fn.lower()
                if lower.endswith(excel_exts) or lower.endswith('.csv') or lower.endswith('.iprd'):
                    all_paths.append(os.path.join(dirpath, fn))

        # --- group by base name (before first dot) ---
        groups = {}
        for p in all_paths:
            base = os.path.basename(p).split('.', 1)[0]
            groups.setdefault(base, []).append(p)

        total_groups = len(groups)
        self.log_message.emit(f"Groups: {total_groups}")
        success, failed = [], []
        produced_merged = []  # abs paths of final merged files

        # --- Pass 1: merge exactly 1 Excel + 1 CSV ---
        for idx, (base, paths) in enumerate(groups.items(), start=1):
            if self._is_cancelled:
                self.status_updated.emit("Cancelled.")
                break

            percent = int((idx / max(total_groups, 1)) * 100)
            self.progress_updated.emit(percent, idx, base)
            self.status_updated.emit(f"Processing: {base}")
            self.log_message.emit(f"[{idx}/{total_groups}] {base}")

            excels = [p for p in paths if p.lower().endswith(excel_exts)]
            csvs   = [p for p in paths if p.lower().endswith('.csv')]
            iprds  = [p for p in paths if p.lower().endswith('.iprd')]

            if len(excels) != 1 or len(csvs) != 1:
                self.log_message.emit("  ⚠️  skip (need exactly 1 Excel + 1 CSV).")
                continue

            xfile, cfile = excels[0], csvs[0]
            merged_fp = os.path.join(self.top_folder, f"{base}_merged.xlsx")
            try:
                self._merge_pair(xfile, cfile, merged_fp)
                produced_merged.append(os.path.abspath(merged_fp))
                success.append((merged_fp, xfile, cfile))
                self.log_message.emit(f"  ✅ merged → {self._fmt_rel(merged_fp)}")

                # move originals (excel/csv/iprd) into backup right away
                try:
                    for src in [xfile, cfile] + iprds:
                        rel_src = os.path.relpath(os.path.abspath(src), os.path.abspath(self.top_folder))
                        # only move things inside top folder
                        if os.path.commonpath([os.path.abspath(src), os.path.abspath(self.top_folder)]) != os.path.abspath(self.top_folder):
                            self.log_message.emit(f"  ⚠️ skip move (outside top): {src}")
                            continue
                        dst = os.path.join(backup_dir, rel_src)
                        self._safe_move(src, dst)
                except Exception as move_err:
                    self.log_message.emit(f"  ⚠️ move originals failed: {move_err}")

            except Exception as e:
                failed.append((base, str(e)))
                self.log_message.emit(f"  ❌ merge failed: {base} → {e}")

        # --- Pass 2: sweep everything else (ALL files) into backup, except final merged in top ---
        self.status_updated.emit("Moving remaining files to backup…")
        self.log_message.emit("— Sweeping remaining files into backup (keep only *_merged.xlsx in root) —")
        merged_set = set(os.path.abspath(p) for p in produced_merged)

        for dirpath, _, filenames in os.walk(self.top_folder):
            abs_dir = os.path.abspath(dirpath)
            if abs_dir.startswith(os.path.abspath(backup_dir)):
                continue
            for fn in filenames:
                p = os.path.join(dirpath, fn)
                ap = os.path.abspath(p)

                # keep final merged files only if they sit in the exact top folder
                if ap in merged_set and abs_dir == os.path.abspath(self.top_folder):
                    continue

                # only move files inside top folder
                if os.path.commonpath([ap, os.path.abspath(self.top_folder)]) != os.path.abspath(self.top_folder):
                    self.log_message.emit(f"⚠️ skip (outside top): {self._fmt_rel(ap)}")
                    continue

                rel = os.path.relpath(ap, self.top_folder)
                dst = os.path.join(backup_dir, rel)
                try:
                    self._safe_move(ap, dst)
                except Exception as e:
                    self.log_message.emit(f"⚠️ move failed: {self._fmt_rel(ap)} → {e}")

        # --- Pass 3: remove now-empty subfolders in top (not touching backup/) ---
        for dirpath, dirnames, _ in os.walk(self.top_folder, topdown=False):
            abs_dir = os.path.abspath(dirpath)
            if abs_dir.startswith(os.path.abspath(backup_dir)):
                continue
            if abs_dir == os.path.abspath(self.top_folder):
                continue
            try:
                if not os.listdir(dirpath):
                    os.rmdir(dirpath)
                    self.log_message.emit(f"🧹 removed empty folder: {self._fmt_rel(dirpath)}")
            except Exception as e:
                self.log_message.emit(f"⚠️ remove dir failed: {self._fmt_rel(dirpath)} → {e}")

        elapsed = time.time() - t0
        self.status_updated.emit(f"Done. Elapsed: {int(elapsed)}s")
        self.log_message.emit(f"✔ Done. Merged: {len(success)}, Failed: {len(failed)}. Elapsed: {int(elapsed)}s")
        self.finished.emit(success, failed)

    def _coerce_excel_number(self, val):
        """Turn common numeric-looking strings into numbers. Return (new_value, number_format or None)."""
        if val is None or isinstance(val, (int, float)):
            return val, None
        if not isinstance(val, str):
            return val, None

        s = val.strip()
        if s == "":
            return "", None

        # negative in parentheses: (123.4) -> -123.4
        m = re.match(r"^\((.+)\)$", s)
        if m:
            s = "-" + m.group(1).strip()

        # percent handling: '12.3%' -> 0.123 and set percent format
        is_percent = s.endswith("%")
        if is_percent:
            s = s[:-1].strip()

        # remove thousands separators (1,234.56 -> 1234.56)
        s = s.replace(",", "")

        # try float
        try:
            num = float(s)
            if is_percent:
                return num / 100.0, "0.00%"
            # optional: choose a general numeric format
            return num, "General"
        except ValueError:
            return val, None

    def _fix_dft_sheet_numbers(self, merged_fp, sheet_name="DFT result"):
        """Open the merged workbook and convert DFT sheet cells to numbers when possible."""
        wb = load_workbook(merged_fp)
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]

        # Heuristic: if first row looks like header (contains alphabetic chars), skip it
        def _looks_like_header(row):
            for c in row:
                v = c.value
                if isinstance(v, str) and any(ch.isalpha() for ch in v):
                    return True
            return False

        start_row = 2 if _looks_like_header(next(ws.iter_rows(min_row=1, max_row=1))) else 1

        changed = 0
        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                new_val, num_fmt = self._coerce_excel_number(cell.value)
                if new_val is not cell.value:
                    cell.value = new_val
                    changed += 1
                if num_fmt:
                    cell.number_format = num_fmt

        wb.save(merged_fp)
        self.log_message.emit(f"  🔢 coerced numbers in '{sheet_name}' (changed ~{changed} cells)")
