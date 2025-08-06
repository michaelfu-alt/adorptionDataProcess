# from PySide6.QtWidgets import (
#     QFileDialog, QDialog, QLabel, QProgressBar, QPushButton, QTextEdit, 
#     QVBoxLayout, QHBoxLayout, QMessageBox
# )

from PySide6.QtCore import QObject, QThread, Signal
from PySide6.QtWidgets import QFileDialog, QMessageBox
import os
from view.process_dialog import ProcessDialog
import os, time
import pandas as pd   # 用于读取Excel
import tempfile
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import matplotlib.pyplot as plt
from PySide6.QtWidgets import QMessageBox


class ImportWorker(QObject):
    progress = Signal(int, int, str)
    finished = Signal(list)
    error = Signal(str)

    def __init__(self, filepaths, model):
        super().__init__()
        self.filepaths = filepaths
        self.model = model
        self._paused = False
        self._cancelled = False
        self.loaded = []

    def pause(self):
        self._paused = True

    def resume(self):
        self._paused = False

    def cancel(self):
        self._cancelled = True

    def run(self):
        print("Worker started")            
        total = len(self.filepaths)

        # 在线程内新建连接，不修改 self.model.conn
        thread_conn = self.model.get_thread_connection()

        try:
            for idx, fp in enumerate(self.filepaths, start=1):
                # print(f"ImportWorker: processing file {idx}/{total}: {fp}")  # DEBUG

                if self._cancelled:
                    break
                while self._paused and not self._cancelled:
                    time.sleep(0.1)
                if self._cancelled:
                    break
                try:
                    filename = os.path.basename(fp)
                    self.progress.emit(idx, total, filename)
                    
                    # 调用 ingest_excel 时传入线程专属连接
                    self.model.ingest_excel(fp, conn=thread_conn)

                    # print(f"ImportWorker: successfully loaded {fp}")  # DEBUG
                    self.loaded.append(filename)
                except Exception as e:
                    print(f"ImportWorker: error loading {fp}: {e}")
                    self.error.emit(f"Failed to import '{filename}':\n{e}")
        finally:
            thread_conn.commit()
            thread_conn.close()  # 用完关闭连接

        self.finished.emit(self.loaded)


class ImportExportManager(QObject):
    # import_started = Signal()
    # import_progress = Signal(int, int, str)  # current, total, filename
    import_error = Signal(str)
    import_finished = Signal(list)  # list of loaded files

    def __init__(self, model, parent=None):
        super().__init__(parent)
        self.model = model
        self.worker = None
        self.thread = None
        self.progress_dialog = None
        self.summary_dialog = None

    def start_import(self, parent_widget):
        filepaths, _ = QFileDialog.getOpenFileNames(
            parent_widget,
            "Select one or more Excel files",
            "",
            "Excel Files (*.xls *.xlsx *.xlsm *.xlsb *.xltx *.xltm);;All Files (*)"
        )
        if not filepaths:
            return

        self.progress_dialog = ProcessDialog(len(filepaths))
        self.progress_dialog.pause_clicked.connect(self.pause_import)
        self.progress_dialog.continue_clicked.connect(self.resume_import)
        self.progress_dialog.end_clicked.connect(self.cancel_import)
        self.progress_dialog.show()

        self.worker = ImportWorker(filepaths, self.model)
        self.worker.progress.connect(self.progress_dialog.update_status)
        self.worker.error.connect(self.on_error)
        self.worker.finished.connect(self.on_finished)

        self.thread = QThread()
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.run)
        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)

        self.thread.start()
        # self.import_started.emit()

    def pause_import(self):
        if self.worker:
            self.worker.pause()

    def resume_import(self):
        if self.worker:
            self.worker.resume()

    def cancel_import(self):
        if self.worker:
            self.worker.cancel()

    def on_error(self, msg):
        self.import_error.emit(msg)

    def on_finished(self, loaded_files):
        if self.progress_dialog:
            self.progress_dialog.close()

        # 这里用主线程重新建立连接，确保安全
        try:
            if self.model.conn:
                self.model.conn.close()
            self.model.conn = self.model.get_main_thread_connection()
        except Exception as e:
            print("Error resetting main DB connection:", e)

        # 弹窗显示导入完成信息    
        summary_text = "\n".join(loaded_files)
        QMessageBox.information(
            self.progress_dialog.parent() or None,
            "Import Summary",
            f"Successfully imported {len(loaded_files)} files:\n{summary_text}"
        )
        self.import_finished.emit(loaded_files)


    def show_summary(parent, loaded_files):
        msg = QMessageBox(parent)
        msg.setWindowTitle("Import Summary")
        msg.setIcon(QMessageBox.Information)
        msg.setText(f"Successfully imported {len(loaded_files)} files:")
        msg.setDetailedText("\n".join(loaded_files))
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec()


EXCEL_CELL_MAP = {
    "样品名称": "B32",
    "吸附质": "B33",
    "样品重量[g]": "B34",
    "吸附质面积[m^2/cc]": "B35",
    "多点BET比表面积[m^2/g]": "D31",
    "多点Langmuir比表面积[m^2/g]": "D32",
    "BJH吸附累积比表面积(d>2[nm])[m^2/g]": "D33",
    "BJH解吸累积比表面积(d>2[nm])[m^2/g]": "D34",
    "t-Plot(吸附)外比表面积[m^2/g]": "D35",
    "t-Plot(吸附)内比表面积[m^2/g]": "D36",
    "αs(吸附)中孔比表面积[m^2/g]": "D37",
    "单点总孔体积[cc/g]": "D38",
    "BJH吸附累积总孔体积(d>2[nm])[cc/g]": "D39",
    "BJH解吸累积总孔体积(d>2[nm])[cc/g]": "D40",
    "单点吸附微孔体积(d<2[nm])[cc/g]": "D41",
    "t-Plot(吸附)微孔体积[cc/g]": "D42",
    "αs(吸附)微孔体积[cc/g]": "D43",
    "单点平均孔半径(面积为BET比表面积)[nm]": "D44",
    "BJH吸附平均孔半径[nm]": "D45",
    "BJH解吸平均孔半径[nm]": "D46",
    "BJH吸附最可几孔径[nm]": "D47",
    "BJH解吸最可几孔径[nm]": "D48",
    "HK最可几孔径[nm]": "D49",
}

# def export_samples_to_excel(path, sample_names, model,
#                             summary_fields=None,
#                             excel_cell_map=None,
#                             parent_widget=None):
#     if summary_fields is None:
#         summary_fields = list(EXCEL_CELL_MAP.keys())
#     if excel_cell_map is None:
#         excel_cell_map = EXCEL_CELL_MAP

#     try:
#         wb = Workbook()
#         default = wb.active
#         wb.remove(default)

#         sheets = []

#         for name in sample_names:
#             info = model.get_sample_info(name)
#             human_name = info.get("Sample Name", name)
#             ws = wb.create_sheet(title=human_name[:31])
#             sheets.append(human_name[:31])

#             ads, des = model.get_adsorption_data(name)

#             # 画吸附脱附曲线
#             figA, axA = plt.subplots(figsize=(4,3))
#             axA.set_title(f"{human_name} Adsorption/Desorption")
#             axA.set_xlabel("P/P₀")
#             axA.set_ylabel("Volume (cc/g)")
#             if ads:
#                 xA, yA = zip(*ads)
#                 axA.plot(xA, yA, color="tab:blue", marker="o", linestyle="-", label="Adsorption")
#             if des:
#                 xD, yD = zip(*des)
#                 axA.plot(xD, yD, color="tab:orange", marker="s", linestyle="--", label="Desorption")
#             axA.grid(False)
#             figA.tight_layout()
#             tmpA = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
#             figA.savefig(tmpA.name, dpi=150)
#             plt.close(figA)
#             imgA = XLImage(tmpA.name)
#             imgA.anchor = "A1"
#             ws.add_image(imgA)

#             # 画PSD图
#             dft_list = model.get_dft_data(name)
#             xs, ys = [], []
#             for rec in dft_list:
#                 dia = rec.get("Pore Diameter(nm)")
#                 psd = rec.get("PSD(total)")
#                 if dia is None or psd is None:
#                     continue
#                 if pd.isna(dia) or pd.isna(psd):
#                     continue
#                 if dia > 10 or psd <= 0.005:
#                     continue
#                 xs.append(dia)
#                 ys.append(psd)
#             figP, axP = plt.subplots(figsize=(4,3))
#             axP.set_title(f"{human_name} PSD")
#             axP.set_xlabel("Diameter (nm)")
#             axP.set_ylabel("PSD (total)")
#             if xs:
#                 axP.plot(xs, ys, color="tab:green", linestyle="-")
#             axP.grid(False)
#             figP.tight_layout()
#             tmpP = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
#             figP.savefig(tmpP.name, dpi=150)
#             plt.close(figP)
#             imgP = XLImage(tmpP.name)
#             imgP.anchor = "F1"
#             ws.add_image(imgP)

#             sample_info = info
#             result_summary = model.get_sample_results(name)
#             info_items = list(sample_info.items())
#             res_items = list(result_summary.items())

#             n_ads = len(ads)
#             n_des = len(des)
#             n_iso = max(n_ads, n_des)

#             psd_rows = []
#             for rec in dft_list:
#                 dia = rec.get("Pore Diameter(nm)")
#                 psd = rec.get("PSD(total)")
#                 if dia is None or psd is None:
#                     continue
#                 if pd.isna(dia) or pd.isna(psd):
#                     continue
#                 if dia > 10 or psd <= 0.005:
#                     continue
#                 psd_rows.append((dia, psd))
#             n_psd = len(psd_rows)

#             n_info = len(info_items)
#             n_res = len(res_items)
#             total_rows = max(n_info, n_res, n_iso, n_psd)

#             header_row = 30
#             headers = [
#                 "Field", "Value",
#                 "Field", "Value",
#                 "P/P₀ (ads)", "V (cc/g ads)",
#                 "P/P₀ (des)", "V (cc/g des)",
#                 "Pore Diameter (nm)", "PSD (total)"
#             ]
#             for col_idx, text in enumerate(headers, start=1):
#                 cell = ws.cell(row=header_row, column=col_idx, value=text)
#                 cell.font = cell.font.copy(bold=True, size=12)
#             for i in range(total_rows):
#                 row_idx = header_row + 1 + i
#                 if i < n_info:
#                     k, v = info_items[i]
#                     ws.cell(row=row_idx, column=1, value=k)
#                     ws.cell(row=row_idx, column=2, value=v)
#                 if i < n_res:
#                     k2, v2 = res_items[i]
#                     ws.cell(row=row_idx, column=3, value=k2)
#                     ws.cell(row=row_idx, column=4, value=v2)
#                 if i < n_ads:
#                     qA, vA = ads[i]
#                     ws.cell(row=row_idx, column=5, value=qA)
#                     ws.cell(row=row_idx, column=6, value=vA)
#                 if i < n_des:
#                     qD, vD = des[i]
#                     ws.cell(row=row_idx, column=7, value=qD)
#                     ws.cell(row=row_idx, column=8, value=vD)
#                 if i < n_psd:
#                     d0, p0 = psd_rows[i]
#                     ws.cell(row=row_idx, column=9, value=d0)
#                     ws.cell(row=row_idx, column=10, value=p0)
#             col_widths = {
#                 1:  20, 2:  25, 3:  20, 4:  25,
#                 5:  12, 6:  15, 7:  12, 8:  15,
#                 9:  15, 10: 15
#             }
#             for col_idx, width in col_widths.items():
#                 ws.column_dimensions[get_column_letter(col_idx)].width = width

#         # 统计汇总sheet
#         all_data = []
#         for sheetname in sheets:
#             ws = wb[sheetname]
#             row_data = []
#             for field in summary_fields:
#                 cell_addr = excel_cell_map.get(field)
#                 if not cell_addr:
#                     continue
#                 value = ws[cell_addr].value
#                 try:
#                     val_clean = str(value).replace(",", "").replace(" ", "")
#                     if val_clean == "" or value is None:
#                         value = None
#                     elif "." in val_clean:
#                         value = float(val_clean)
#                     else:
#                         value = int(val_clean)
#                 except Exception:
#                     pass
#                 row_data.append(value)
#             all_data.append(row_data)

#         df = pd.DataFrame(all_data, columns=summary_fields)

#         ws_sum = wb.create_sheet(title="统计汇总", index=0)

#         # 写表头
#         for col_idx, field in enumerate(summary_fields, 1):
#             ws_sum.cell(row=1, column=col_idx, value=field)

#         # 写详细数据
#         for row_idx, row in enumerate(all_data, 2):
#             for col_idx, value in enumerate(row, 1):
#                 ws_sum.cell(row=row_idx, column=col_idx, value=value)

#         # 写统计量
#         stat_funcs = [
#             ("均值", lambda x: np.nanmean(x)),
#             ("最大", lambda x: np.nanmax(x)),
#             ("最小", lambda x: np.nanmin(x)),
#             ("标准差", lambda x: np.nanstd(x)),
#         ]

#         n_data = len(all_data)
#         for i, (stat_name, func) in enumerate(stat_funcs, 1):
#             out_row = n_data + 1 + i
#             for col_idx, field in enumerate(summary_fields, 1):
#                 if col_idx == 1:
#                     ws_sum.cell(row=out_row, column=col_idx, value=stat_name)
#                 else:
#                     vals = pd.to_numeric(df[field], errors='coerce').values
#                     try:
#                         stat = func(vals)
#                         stat = "" if np.isnan(stat) else round(stat, 3)
#                     except Exception:
#                         stat = ""
#                     ws_sum.cell(row=out_row, column=col_idx, value=stat)

#         # 美化
#         for col_idx in range(1, len(summary_fields) + 1):
#             ws_sum.column_dimensions[get_column_letter(col_idx)].width = 22
#         for row in range(1, ws_sum.max_row + 1):
#             ws_sum.row_dimensions[row].height = 26

#         big_font = Font(name='微软雅黑', size=14)
#         align_center = Alignment(horizontal='center', vertical='center')
#         for row in ws_sum.iter_rows(min_row=1, max_row=ws_sum.max_row, min_col=1, max_col=len(summary_fields)):
#             for cell in row:
#                 cell.font = big_font
#                 cell.alignment = align_center

#         header_font = Font(name='微软雅黑', size=15, bold=True)
#         for cell in ws_sum[1]:
#             cell.font = header_font

#         n_stat = len(stat_funcs)
#         for row_i in range(n_data + 2, n_data + 2 + n_stat):
#             for cell in ws_sum[row_i]:
#                 cell.font = Font(name='微软雅黑', size=14, bold=True)

#         wb.save(path)
#     except Exception as e:
#         import traceback
#         print(traceback.format_exc())
#         if parent_widget:
#             QMessageBox.critical(parent_widget, "Export Error", f"Failed to export samples:\n{e}")
#         else:
#             print(f"Failed to export samples:\n{e}")

class SampleExporter:
    EXCEL_CELL_MAP = {
        "样品名称": "B32",
        "吸附质": "B33",
        "样品重量[g]": "B34",
        "吸附质面积[m^2/cc]": "B35",
        "多点BET比表面积[m^2/g]": "D31",
        "多点Langmuir比表面积[m^2/g]": "D32",
        "BJH吸附累积比表面积(d>2[nm])[m^2/g]": "D33",
        "BJH解吸累积比表面积(d>2[nm])[m^2/g]": "D34",
        "t-Plot(吸附)外比表面积[m^2/g]": "D35",
        "t-Plot(吸附)内比表面积[m^2/g]": "D36",
        "αs(吸附)中孔比表面积[m^2/g]": "D37",
        "单点总孔体积[cc/g]": "D38",
        "BJH吸附累积总孔体积(d>2[nm])[cc/g]": "D39",
        "BJH解吸累积总孔体积(d>2[nm])[cc/g]": "D40",
        "单点吸附微孔体积(d<2[nm])[cc/g]": "D41",
        "t-Plot(吸附)微孔体积[cc/g]": "D42",
        "αs(吸附)微孔体积[cc/g]": "D43",
        "单点平均孔半径(面积为BET比表面积)[nm]": "D44",
        "BJH吸附平均孔半径[nm]": "D45",
        "BJH解吸平均孔半径[nm]": "D46",
        "BJH吸附最可几孔径[nm]": "D47",
        "BJH解吸最可几孔径[nm]": "D48",
        "HK最可几孔径[nm]": "D49",
    }

    def __init__(self, model, parent_widget=None):
        """
        :param model: 数据模型，必须实现 get_sample_info, get_adsorption_data, get_dft_data, get_sample_results 方法
        :param parent_widget: 用于消息框的父窗口，PySide6 QWidget 或 None
        """
        self.model = model
        self.parent_widget = parent_widget

    def export(self, path, sample_names, summary_fields=None, excel_cell_map=None):
        """
        导出样品到Excel文件
        :param path: 保存路径
        :param sample_names: 样品内部唯一标识列表
        :param summary_fields: 统计汇总字段列表，默认取 EXCEL_CELL_MAP 所有字段
        :param excel_cell_map: 字段到单元格映射，默认 EXCEL_CELL_MAP
        """
        if summary_fields is None:
            summary_fields = list(self.EXCEL_CELL_MAP.keys())
        if excel_cell_map is None:
            excel_cell_map = self.EXCEL_CELL_MAP

        try:
            wb = Workbook()
            default = wb.active
            wb.remove(default)
            sheets = []

            for name in sample_names:
                self._write_sample_sheet(wb, name, sheets)

            self._write_summary_sheet(wb, sheets, summary_fields, excel_cell_map)
            wb.save(path)

            QMessageBox.information(self.parent_widget, "导出完成", f"成功导出 {len(sample_names)} 个样品到：\n{path}")

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            QMessageBox.critical(self.parent_widget, "导出错误", f"导出失败:\n{e}")
    



    def _write_sample_sheet(self, wb, sample_name, sheets):
        info = self.model.get_export_sample_info(sample_name)
        print(info)
        print(sample_name)
        human_name = info.get("Sample Name", sample_name)
        print(f"Sample: {human_name}")
        print(f"Writing Excel sheet for sample: {human_name}")
        ws = wb.create_sheet(title=human_name[:31])
        sheets.append(human_name[:31])

        ads, des = self.model.get_adsorption_data(sample_name)
        print(f"Adsorption points count: {len(ads)}")
        print(f"Desorption points count: {len(des)}")

        # 画吸附脱附曲线
        figA, axA = plt.subplots(figsize=(4,3))
        axA.set_title(f"{human_name} Adsorption/Desorption")
        axA.set_xlabel("P/P₀")
        axA.set_ylabel("Volume (cc/g)")
        if ads:
            xA, yA = zip(*ads)
            axA.plot(xA, yA, color="tab:blue", marker="o", linestyle="-", label="Adsorption")
        if des:
            xD, yD = zip(*des)
            axA.plot(xD, yD, color="tab:orange", marker="s", linestyle="--", label="Desorption")
        axA.grid(False)
        figA.tight_layout()
        tmpA = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        figA.savefig(tmpA.name, dpi=150)
        plt.close(figA)
        imgA = XLImage(tmpA.name)
        imgA.anchor = "A1"
        ws.add_image(imgA)

        # 画PSD图
        dft_list = self.model.get_dft_data(sample_name)
        xs, ys = [], []
        for rec in dft_list:
            dia = rec.get("Pore Diameter(nm)")
            psd = rec.get("PSD(total)")
            if dia is None or psd is None:
                continue
            if pd.isna(dia) or pd.isna(psd):
                continue
            if dia > 10 or psd <= 0.005:
                continue
            xs.append(dia)
            ys.append(psd)
        figP, axP = plt.subplots(figsize=(4,3))
        axP.set_title(f"{human_name} PSD")
        axP.set_xlabel("Diameter (nm)")
        axP.set_ylabel("PSD (total)")
        if xs:
            axP.plot(xs, ys, color="tab:green", linestyle="-")
        axP.grid(False)
        figP.tight_layout()
        tmpP = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        figP.savefig(tmpP.name, dpi=150)
        plt.close(figP)
        imgP = XLImage(tmpP.name)
        imgP.anchor = "F1"
        ws.add_image(imgP)

        sample_info = info
        result_summary = self.model.get_sample_results(sample_name)
        info_items = list(sample_info.items())
        res_items = list(result_summary.items())

        n_ads = len(ads)
        n_des = len(des)
        n_iso = max(n_ads, n_des)

        psd_rows = []
        for rec in dft_list:
            dia = rec.get("Pore Diameter(nm)")
            psd = rec.get("PSD(total)")
            if dia is None or psd is None:
                continue
            if pd.isna(dia) or pd.isna(psd):
                continue
            if dia > 10 or psd <= 0.005:
                continue
            psd_rows.append((dia, psd))
        n_psd = len(psd_rows)

        n_info = len(info_items)
        n_res = len(res_items)
        total_rows = max(n_info, n_res, n_iso, n_psd)

        header_row = 30
        headers = [
            "Field", "Value",
            "Field", "Value",
            "P/P₀ (ads)", "V (cc/g ads)",
            "P/P₀ (des)", "V (cc/g des)",
            "Pore Diameter (nm)", "PSD (total)"
        ]
        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=text)
            cell.font = cell.font.copy(bold=True, size=12)
        for i in range(total_rows):
            row_idx = header_row + 1 + i
            if i < n_info:
                k, v = info_items[i]
                ws.cell(row=row_idx, column=1, value=k)
                ws.cell(row=row_idx, column=2, value=v)
            if i < n_res:
                k2, v2 = res_items[i]
                ws.cell(row=row_idx, column=3, value=k2)
                ws.cell(row=row_idx, column=4, value=v2)
            if i < n_ads:
                qA, vA = ads[i]
                ws.cell(row=row_idx, column=5, value=qA)
                ws.cell(row=row_idx, column=6, value=vA)
            if i < n_des:
                qD, vD = des[i]
                ws.cell(row=row_idx, column=7, value=qD)
                ws.cell(row=row_idx, column=8, value=vD)
            if i < n_psd:
                d0, p0 = psd_rows[i]
                ws.cell(row=row_idx, column=9, value=d0)
                ws.cell(row=row_idx, column=10, value=p0)
        col_widths = {
            1:  20, 2:  25, 3:  20, 4:  25,
            5:  12, 6:  15, 7:  12, 8:  15,
            9:  15, 10: 15
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_summary_sheet(self, wb, sheets, summary_fields, excel_cell_map):
        all_data = []
        for sheetname in sheets:
            ws = wb[sheetname]
            row_data = []
            for field in summary_fields:
                cell_addr = excel_cell_map.get(field)
                if not cell_addr:
                    continue
                value = ws[cell_addr].value
                try:
                    val_clean = str(value).replace(",", "").replace(" ", "")
                    if val_clean == "" or value is None:
                        value = None
                    elif "." in val_clean:
                        value = float(val_clean)
                    else:
                        value = int(value)
                except Exception:
                    pass
                row_data.append(value)
            all_data.append(row_data)

        df = pd.DataFrame(all_data, columns=summary_fields)

        ws_sum = wb.create_sheet(title="统计汇总", index=0)

        # 写表头
        for col_idx, field in enumerate(summary_fields, 1):
            ws_sum.cell(row=1, column=col_idx, value=field)

        # 写详细数据
        for row_idx, row in enumerate(all_data, 2):
            for col_idx, value in enumerate(row, 1):
                ws_sum.cell(row=row_idx, column=col_idx, value=value)

        # 写统计量
        stat_funcs = [
            ("均值", lambda x: np.nanmean(x)),
            ("最大", lambda x: np.nanmax(x)),
            ("最小", lambda x: np.nanmin(x)),
            ("标准差", lambda x: np.nanstd(x)),
        ]

        n_data = len(all_data)
        for i, (stat_name, func) in enumerate(stat_funcs, 1):
            out_row = n_data + 1 + i
            for col_idx, field in enumerate(summary_fields, 1):
                if col_idx == 1:
                    ws_sum.cell(row=out_row, column=col_idx, value=stat_name)
                else:
                    vals = pd.to_numeric(df[field], errors='coerce').values
                    try:
                        stat = func(vals)
                        stat = "" if np.isnan(stat) else round(stat, 3)
                    except Exception:
                        stat = ""
                    ws_sum.cell(row=out_row, column=col_idx, value=stat)

        # 美化
        for col_idx in range(1, len(summary_fields) + 1):
            ws_sum.column_dimensions[get_column_letter(col_idx)].width = 22
        for row in range(1, ws_sum.max_row + 1):
            ws_sum.row_dimensions[row].height = 26

        big_font = Font(name='微软雅黑', size=14)
        align_center = Alignment(horizontal='center', vertical='center')
        for row in ws_sum.iter_rows(min_row=1, max_row=ws_sum.max_row, min_col=1, max_col=len(summary_fields)):
            for cell in row:
                cell.font = big_font
                cell.alignment = align_center

        header_font = Font(name='微软雅黑', size=15, bold=True)
        for cell in ws_sum[1]:
            cell.font = header_font

        n_stat = len(stat_funcs)
        for row_i in range(n_data + 2, n_data + 2 + n_stat):
            for cell in ws_sum[row_i]:
                cell.font = Font(name='微软雅黑', size=14, bold=True)